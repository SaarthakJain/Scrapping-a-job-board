import os
import urllib, urllib2
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import read
import time
import sys

# Creating a 2-D list for storing the result
output = []
output.append([])
output.append([])
output.append([])
output.append([])
output.append([])
# Creating a list to store the funds
funding = []

# This function scraped the url and stores the record in associated lists
def scrapeInfo(url):
    # Make a Selenium webdriver
    browser = webdriver.Firefox()
    # Run the url in the browser
    browser.get(url)
    elem = ""
    html_page = browser.page_source
    # Converting the html source to utf-8 format
    string = html_page.encode('utf-8')
    # Generating BeautifulSoup text to extract the contents of tags
    soup = BeautifulSoup(string)
    elem = soup.find("div",class_="count-box").contents
    k = str(elem[1])
    #print k
    p = k.split('>')
    #print p
    k = p[1].split('<')
    c = int(k[0])
    # c stores the number of startups found
    for i in range (0,c//10):
        # Scroll to page end
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # Wait for page to load further
        # To be on safer side because of network connection the time limit is set to 5 sec
        time.sleep(5)
    html_page = browser.page_source
    string = html_page.encode('utf-8')
    soup = BeautifulSoup(string)    
    browser.quit()
    # Exit the browser and now only work on the extracted cource code

    # Find Company Name
    x = soup.find_all("div", class_="startup-row section")
    for company in x:
        output[0].append(company.get('data-startup_name'))
    # Find Compnay URL
    for a in soup.find_all("a", class_="website-link"):
        output[1].append(a.get('href'))
    # Find no. of employees
    for employees in soup.find_all("div", class_="tag employees"):
        output[2].append(''.join(employees.find_all(text=True)))
    # Find locations
    for locations in soup.find_all("div", class_="tag locations tiptip"):
        output[3].append(''.join(locations.find_all(text=True)))
    # Find areas
    for markets in soup.find_all("div", class_="tag markets"):
        output[4].append(''.join(markets.find_all(text=True)))
    # Find company URL on angel.co to extract the funding
    x = soup.find_all("div", class_="company-name")
    for div in x:
        req = urllib2.Request(div.a['href'])
        res = urllib2.urlopen(req)
        new_soup = BeautifulSoup(res.read())
        funds = []
        for fund in new_soup.find_all("div", "raised"):
            funds.append(''.join(fund.find_all(text=True)))
            #print funds
        funding.append(funds)
    #print funding

# Used to write the text to excel file
def writeToExcel():
    # If file exist then open the file and create new sheet after existing ones
    if os.path.exists('result.xlsx'):
        wb = load_workbook('result.xlsx')
        c = 0
        for sheet in wb:
            c = c + 1
        ws = wb.create_sheet(c + 1)
        ws.title = "Sheet_" + str(c+1)
    else:
        # Create a new workbook and new sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet_1"
    # Storing values in each cell
    ws.cell('%s%s'%('A', 1)).value = "Company Name"
    ws.cell('%s%s'%('B', 1)).value = "Company URL"
    ws.cell('%s%s'%('C', 1)).value = "Employee Count"
    ws.cell('%s%s'%('D', 1)).value = "Location"
    ws.cell('%s%s'%('E', 1)).value = "Functional Area"
    ws.cell('%s%s'%('F', 1)).value = "Funding"

    # Creating a list to be used for refenrencing the cells
    pos = ['A','B','C','D','E']
    for i,result in enumerate(output):
        for j,value in enumerate(result):
            ws.cell('%s%s'%(pos[i], j+2)).value = result[j].strip('/n')

    r = 2
    for one_company in funding:
        text=""
        for value in one_company:
            text = text + value.strip('\n') + ","
        #print text[:-1]
        ws.cell('%s%s'%('F', r)).value = text[:-1]
        r = r + 1
        
    wb.save('result.xlsx')
    # Finally saving the file

if __name__ == "__main__":
    sys.setrecursionlimit(5000)
    url = read.readInput('newfile.yml')
    print url
    scrapeInfo(url)
    writeToExcel()
